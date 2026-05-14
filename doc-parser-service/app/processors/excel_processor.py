import time
import uuid
from pathlib import Path
from loguru import logger
from openpyxl import load_workbook

from app.core.base_parser import BaseParser
from app.models.response import (
    ParseResult,
    DocumentMetadata,
    LogicalSegment,
    PageDebugInfo,
    LayoutBlock,
    BlockType,
    BoundingBox,
)


class ExcelProcessor(BaseParser):
    def supports(self, file_extension: str, strategy: str) -> bool:
        return file_extension in ("xlsx", "xls")

    def order(self) -> int:
        return 10

    def parse(self, file_path: Path, strategy: str, language: str, **kwargs) -> ParseResult:
        start = time.time()
        try:
            wb = load_workbook(str(file_path), read_only=True, data_only=True)

            segments: list[LogicalSegment] = []
            blocks: list[LayoutBlock] = []
            order = 0

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                table_html = self._sheet_to_html(ws, sheet_name)

                block_id = str(uuid.uuid4())
                blocks.append(LayoutBlock(
                    id=block_id,
                    type=BlockType.TABLE,
                    bbox=BoundingBox(x0=0, y0=0, x1=100, y1=100),
                    html=table_html,
                    page_num=1,
                    confidence=1.0,
                    attributes={"sheet_name": sheet_name},
                ))

                segments.append(LogicalSegment(
                    id=block_id,
                    type="table",
                    content=table_html,
                    start_order=order,
                    end_order=order,
                    metadata={"sheet_name": sheet_name},
                ))
                order += 1

            wb.close()

            page_info = PageDebugInfo(
                page_num=1,
                column_count=1,
                blocks=blocks,
            )

            elapsed = int((time.time() - start) * 1000)
            return ParseResult(
                success=True,
                segments=segments,
                pages=[page_info],
                document_metadata=DocumentMetadata(
                    filename=file_path.name,
                    file_type=file_path.suffix.lstrip("."),
                    page_count=1,
                    language=language,
                    parse_strategy=strategy,
                    parser_engine="openpyxl",
                ),
                processing_time_ms=elapsed,
            )
        except Exception as e:
            logger.exception("Excel parse failed")
            elapsed = int((time.time() - start) * 1000)
            return ParseResult(success=False, error_message=str(e), processing_time_ms=elapsed)

    def _sheet_to_html(self, ws, sheet_name: str) -> str:
        html = f'<table data-sheet="{sheet_name}">'

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            html += "</table>"
            return html

        for row_idx, row in enumerate(rows):
            html += "<tr>"
            for cell_value in row:
                text = str(cell_value) if cell_value is not None else ""
                tag = "th" if row_idx == 0 else "td"
                html += f"<{tag}>{text}</{tag}>"
            html += "</tr>"

        html += "</table>"
        return html
